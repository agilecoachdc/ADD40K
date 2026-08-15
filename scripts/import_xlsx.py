#!/usr/bin/env python3
"""Import ponctuel des fiches ADD40K depuis Excel vers l'app web.

Génère :
  - src/shared/reference-data.ts   (catalogue de règles, dédupliqué depuis
                                     la feuille `listes`, identique dans les
                                     8 fichiers sources)
  - scripts/characters.seed.json   (un objet Character par personnage,
                                     extrait de la feuille `données` de
                                     chaque fichier)
  - scripts/import-report.md       (écarts et anomalies détectés, à lire
                                     avant de considérer l'import "propre")

Usage : python3 scripts/import_xlsx.py
Source : Mon Drive/ADD40K/Fiches persos/*.xlsx (8 fichiers, cf. plan
lively-rolling-comet.md — dossier Backgrounds volontairement ignoré).

Ce script est *déclaratif sur les données Excel connues au 15/08/2026* :
si de nouvelles fiches sont ajoutées ou modifiées dans le dossier Drive, le
relancer régénère les 3 fichiers de sortie ci-dessus.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = Path(
    "/Users/charles/Library/CloudStorage/GoogleDrive-charlesdeclarens@gmail.com/"
    "Mon Drive/ADD40K/Fiches persos"
)
REFERENCE_FILE = SOURCE_DIR / "Stern Tack.xlsx"  # catalogue `listes` identique dans les 8 fichiers

ATTRS = ["FO", "VIT", "DEX", "REF", "PER", "COM", "INT", "VOL"]

# ---------------------------------------------------------------------------
# Table des bonus raciaux, extraite des formules IF de la feuille `données`
# (cellules D6,D8,D10,D12,D14,D16,D18,D20 — une par attribut, une branche par
# race) de Stern Tack.xlsx, identiques dans les 8 fichiers. Cross-vérifiée
# avec "Règles ADD40K - V0.2.docx" : la plupart des races correspondent
# exactement au texte des règles, à deux exceptions près (laissées telles
# que jouées dans le classeur, signalées dans le rapport d'import) :
#   - gnome : le docx indique Intelligence +1, le classeur ne l'applique pas
#   - orc   : le docx indique Intelligence -1, le classeur ne l'applique pas
RACE_ATTRIBUTE_BONUS = {
    "eldar":    {"FO": 0, "VIT": -1, "DEX": 1, "REF": 0, "PER": 1, "COM": 0, "INT": 1, "VOL": 1},
    "rohirim":  {"FO": 0, "VIT": 1, "DEX": 1, "REF": -1, "PER": 0, "COM": 0, "INT": 1, "VOL": 1},
    "gith":     {"FO": 0, "VIT": 0, "DEX": 0, "REF": 0, "PER": 0, "COM": 0, "INT": 1, "VOL": 1},
    "rakshasa": {"FO": 0, "VIT": 0, "DEX": 0, "REF": 1, "PER": 0, "COM": 1, "INT": 0, "VOL": 0},
    "hobbit":   {"FO": 0, "VIT": -1, "DEX": 1, "REF": 1, "PER": 1, "COM": 1, "INT": 0, "VOL": 0},
    "orc":      {"FO": 1, "VIT": 1, "DEX": 0, "REF": 0, "PER": -1, "COM": 0, "INT": 0, "VOL": 0},
    "gnome":    {"FO": -1, "VIT": 0, "DEX": 1, "REF": 0, "PER": 0, "COM": 0, "INT": 0, "VOL": 0},
    "humain":   {"FO": 0, "VIT": 0, "DEX": 0, "REF": 0, "PER": 0, "COM": 0, "INT": 0, "VOL": 0},
}
# Taille (données!B24), par race.
RACE_TAILLE_BONUS = {
    "eldar": 0, "rohirim": -1, "gith": 0, "rakshasa": 0,
    "hobbit": -1, "orc": 1, "gnome": -1, "humain": 0,
}
# Points de compétence de départ (données!F2). "humain" est absent de la
# formule IF du classeur (branche manquante, pas une erreur de calcul jouée
# — aucun perso humain n'existait avant Johnny/Jonas) -> valeur du docx (20)
# utilisée pour compléter.
RACE_SKILL_POINTS = {
    "eldar": 10, "rohirim": 15, "gith": 15, "rakshasa": 15,
    "hobbit": 15, "orc": 20, "gnome": 20, "humain": 20,
}
def hp_max(attribute_scores: dict, tech_bonus: dict, race: str) -> int:
    """Réplique calc-engine.getHpMax en Python (PV = (VIT_total + taille) x 5 + 30)."""
    vit_total = (
        (attribute_scores.get("VIT") or 0)
        + RACE_ATTRIBUTE_BONUS.get(race, {}).get("VIT", 0)
        + (tech_bonus.get("VIT") or 0)
    )
    taille = RACE_TAILLE_BONUS.get(race, 0)
    return round((vit_total + taille) * 5 + 30)


def psp_max(attribute_scores: dict, tech_bonus: dict, race: str) -> int:
    """Réplique calc-engine.getPspMax en Python (PSP = VOL_total x 5 + 30)."""
    vol_total = (
        (attribute_scores.get("VOL") or 0)
        + RACE_ATTRIBUTE_BONUS.get(race, {}).get("VOL", 0)
        + (tech_bonus.get("VOL") or 0)
    )
    return round(vol_total * 5 + 30)


RACE_LABELS = {
    "eldar": "Eldar", "rohirim": "Rohirim", "gith": "Gith", "rakshasa": "Rakshasa",
    "hobbit": "Hobbit", "orc": "Orc", "gnome": "Gnome", "humain": "Humain",
}

# Table des localisations (char_sheet!AT131, identique sur toutes les fiches) —
# règle de jeu statique, pas une donnée par personnage.
LOCALISATIONS = "1-2 jambe g.\n3-4 jambe d.\n5,6,7 torse\n8 bras g.\n9 bras d.\n10 tête"

report_lines: list[str] = []


def note(line: str) -> None:
    report_lines.append(line)
    print(line)


# ---------------------------------------------------------------------------
# Extraction du catalogue (feuille `listes`)
# ---------------------------------------------------------------------------

def extract_skill_cost_table(listes) -> dict[int, int]:
    table = {0: 0}  # listes!A14 est vide (score implicite 0), B14=0
    row = 15
    while listes[f"A{row}"].value is not None:
        score = int(listes[f"A{row}"].value)
        cost = int(listes[f"B{row}"].value)
        table[score] = cost
        row += 1
    return table


ATTR_IN_PARENS = re.compile(r"\((FO|VIT|DEX|REF|PER|COM|INT|VOL)\)")


def extract_skills_catalog(listes) -> list[dict]:
    skills = []
    row = 14
    while listes[f"D{row}"].value is not None:
        name = str(listes[f"D{row}"].value).strip()
        match = ATTR_IN_PARENS.search(name)
        skills.append({"name": name, "attribute": match.group(1) if match else None})
        row += 1
    return skills


def extract_psy_powers_catalog(listes) -> list[dict]:
    powers = []
    row = 14
    while True:
        name = listes[f"F{row}"].value
        if name is None:
            row += 1
            if row > 32:
                break
            continue
        if str(name).strip() == "Avantages et inconvénients":
            break
        discipline = listes[f"G{row}"].value
        powers.append({"name": str(name).strip(), "discipline": str(discipline).strip() if discipline else ""})
        row += 1
    return powers


def extract_advantages_catalog(listes) -> list[dict]:
    """F33:G112 — libellé + valeur signée. La ligne 112 ('Dans les nuages: -30')
    est la valeur CORRECTE du catalogue ; le bug du classeur (cf. plan) est
    dans la feuille `données` de chaque personnage (mauvaise cellule référencée
    par la formule K36), pas ici. En dérivant systématiquement la valeur des
    avantages depuis ce catalogue plutôt que depuis la cellule mise en cache
    par personnage, le bug ne se propage pas dans l'import."""
    advantages = []
    row = 33
    while listes[f"F{row}"].value is not None:
        label = str(listes[f"F{row}"].value).strip()
        value = listes[f"G{row}"].value
        advantages.append({"label": label, "value": float(value) if value is not None else 0})
        row += 1
    return advantages


def extract_weapons_catalog(listes) -> list[dict]:
    """I/J/K/L/R, alignées ligne à ligne (name, damage, price, RA, type).
    N/O et Q sont des copies redondantes utilisées par d'autres VLOOKUP dans
    le classeur — ignorées ici, une seule entrée par arme suffit."""
    weapons = []
    row = 14
    while listes[f"I{row}"].value is not None:
        name = str(listes[f"I{row}"].value).strip()
        weapons.append({
            "name": name,
            "damage": listes[f"J{row}"].value,
            "price": listes[f"K{row}"].value,
            "ra": listes[f"L{row}"].value,
            "type": listes[f"R{row}"].value or listes[f"M{row}"].value,
        })
        row += 1
    return weapons


def extract_armor_catalog(listes) -> list[dict]:
    """T/U/X/AA/AD, alignées ligne à ligne (name, VP tête/bras/torse/jambes)."""
    armor = []
    row = 14
    while listes[f"T{row}"].value is not None:
        armor.append({
            "name": str(listes[f"T{row}"].value).strip(),
            "vpTete": listes[f"U{row}"].value or 0,
            "vpBras": listes[f"X{row}"].value or 0,
            "vpTorse": listes[f"AA{row}"].value or 0,
            "vpJambes": listes[f"AD{row}"].value or 0,
        })
        row += 1
    return armor


def build_reference_data() -> dict:
    wb = openpyxl.load_workbook(REFERENCE_FILE, data_only=True)
    listes = wb["listes"]

    races = [
        {
            "race": race,
            "label": RACE_LABELS[race],
            "attributeBonus": bonus,
            "tailleBonus": RACE_TAILLE_BONUS[race],
            "skillPoints": RACE_SKILL_POINTS[race],
        }
        for race, bonus in RACE_ATTRIBUTE_BONUS.items()
    ]

    return {
        "races": races,
        "skillCostTable": extract_skill_cost_table(listes),
        "skills": extract_skills_catalog(listes),
        "weapons": extract_weapons_catalog(listes),
        "armor": extract_armor_catalog(listes),
        "psyPowers": extract_psy_powers_catalog(listes),
        "advantages": extract_advantages_catalog(listes),
        "localisations": LOCALISATIONS,
    }


# ---------------------------------------------------------------------------
# Extraction d'un personnage (feuilles `données` + `char_sheet`)
# ---------------------------------------------------------------------------

def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug or "personnage"


# Mapping joueur -> personnage connu depuis les noms de fichiers du dossier
# Backgrounds ("Background Karun (perso jo).docx", "Background Stern Tack
# (perso Charles).gdoc", "Background Frigg (perso Jack).jpeg"). Pour les
# personnages sans indice trouvé, le username est un slug du nom du
# personnage — À CORRIGER PAR LE MJ après import (cf. docs/API_REFERENCE.md
# et le rapport d'import).
KNOWN_OWNERS = {
    "Karun": "jo",
    "Stern Tack": "charles",
    "Frigg": "jack",
}


def cell(ws, coord):
    v = ws[coord].value
    return v


def extract_equipment(char_sheet) -> list[dict]:
    items = []
    for row in range(9, 52):
        v = char_sheet[f"DD{row}"].value
        if v and str(v).strip() and str(v).strip() not in (
            "Equipement", "Neuromat",  # 'Neuromat' est un en-tête de section, pas un item
        ):
            items.append({"label": str(v).strip()})
    return items


def extract_character(path: Path, advantages_catalog: list[dict]) -> tuple[dict, list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(path, data_only=True)
    dn = wb["données"]
    cs = wb["char_sheet"]

    raw_name = cell(cs, "N6")
    name = raw_name if raw_name and "___" not in str(raw_name) else path.stem
    name = str(name).strip()

    race = str(dn["C2"].value or "").strip().lower()
    if race not in RACE_ATTRIBUTE_BONUS:
        warnings.append(
            f"race '{race}' absente des 8 races gérées par le calculateur — "
            "bonus racial et points de compétence de départ à 0, à corriger manuellement."
        )

    attribute_scores = {}
    attribute_tech_bonus = {}
    for attr, row in zip(ATTRS, [6, 8, 10, 12, 14, 16, 18, 20]):
        attribute_scores[attr] = dn[f"B{row}"].value or 0
        # données!F{row} = bonus "tech" manuel (cyberware, etc.), en plus du
        # bonus racial (colonne D). Rare mais bien réel : ex. Conrad Lingus a
        # VIT F8=+2 et VOL F20=+2, non listés nulle part ailleurs sur la fiche.
        tech = dn[f"F{row}"].value
        if tech:
            attribute_tech_bonus[attr] = tech

    skills = []
    for row in range(3, 15):
        n = dn[f"J{row}"].value
        if n:
            skills.append({"name": str(n).strip(), "score": dn[f"K{row}"].value or 0})

    # Armes lues depuis char_sheet plutôt que données : char_sheet peut
    # contenir des surcharges manuelles (ex. Widowmaker) qui divergent de la
    # formule données — char_sheet est ce que le joueur voit et utilise
    # réellement à table. 5 emplacements, alignés avec données!O3:O7 aux
    # lignes char_sheet 83/86/89/92/95 (cf. investigation "score de base").
    weapons = []
    for row in (83, 86, 89, 92, 95):
        n = cs[f"J{row}"].value
        if n:
            weapons.append({
                "name": str(n).strip(),
                "type": cs[f"AO{row}"].value or "Fire",
                "damage": cs[f"AJ{row}"].value or 0,
                "ra": cs[f"AF{row}"].value or 0,
                # Score de base pour toucher (char_sheet!AY) : présent pour la
                # plupart des persos mais absent sur certaines fiches (Frigg,
                # Jonas) — laissé à 0 dans ce cas, éditable dans l'app.
                "baseScore": cs[f"AY{row}"].value or 0,
            })

    armor = []
    for row in range(13, 18):
        n = dn[f"O{row}"].value
        if n:
            armor.append({
                "name": str(n).strip(),
                "vpTete": dn[f"P{row}"].value or 0,
                "vpBras": dn[f"Q{row}"].value or 0,
                "vpTorse": dn[f"R{row}"].value or 0,
                "vpJambes": dn[f"S{row}"].value or 0,
                # Toutes les armures listées sont considérées équipées par
                # défaut à l'import (c'était déjà la seule liste connue du
                # personnage) — à décocher dans l'app selon la situation.
                "active": True,
            })

    psy_powers = []
    for row in range(19, 24):
        n = dn[f"J{row}"].value
        if n:
            psy_powers.append({
                "name": str(n).strip(),
                "score": dn[f"K{row}"].value or 0,
                "discipline": str(dn[f"M{row}"].value or "").strip(),
            })

    catalog_by_label = {a["label"]: a["value"] for a in advantages_catalog}
    advantages = []
    for row in range(28, 40):
        label = dn[f"J{row}"].value
        if label:
            label = str(label).strip()
            if label in catalog_by_label:
                value = catalog_by_label[label]
            else:
                value = dn[f"K{row}"].value or 0
                warnings.append(
                    f"avantage '{label}' introuvable dans le catalogue — valeur mise en cache "
                    f"dans le classeur reprise telle quelle ({value})."
                )
            advantages.append({"label": label, "value": value})

    points_depart = dn["H23"].value or 0
    xp = dn["H24"].value
    if xp is None or isinstance(xp, str):
        warnings.append(f"XP (H24) invalide dans le classeur ({xp!r}) — mis à 0.")
        xp = 0

    character = {
        "id": slugify(name),
        "ownerUsername": KNOWN_OWNERS.get(name, slugify(name)),
        "name": name,
        "age": cs["AC6"].value if isinstance(cs["AC6"].value, (int, float)) else None,
        "heightM": None,
        "weightLabel": str(cs["AD9"].value or ""),
        "race": race,
        # char_sheet!I9 est le LIBELLÉ statique "Loyauté:" (pas une donnée par
        # personnage) ; la valeur réelle est en O9 (ex. "Alliance"). Il n'y a
        # pas de cellule "faction" distincte dans le classeur — un seul champ
        # existe, laissé vide ici et alimenté via `loyaute`.
        "faction": "",
        "fonction": str(cs["AS9"].value or ""),
        "loyaute": str(cs["O9"].value or ""),
        "portraitUrl": None,
        "attributeScores": attribute_scores,
        "attributeTechBonus": attribute_tech_bonus,
        "tailleModifier": RACE_TAILLE_BONUS.get(race, 0),
        # PV/PSP "actuel" est une nouveauté de l'app (l'Excel n'affichait que
        # max/max, cf. char_sheet!BH8="='données'!E24" recopié en BN8) —
        # initialisé au max calculé, donc personnage "en pleine forme" à l'import.
        "hpCurrent": hp_max(attribute_scores, attribute_tech_bonus, race),
        "pspCurrent": psp_max(attribute_scores, attribute_tech_bonus, race),
        "skills": skills,
        "psyPowers": psy_powers,
        "weapons": weapons,
        "armor": armor,
        "advantages": advantages,
        "equipment": extract_equipment(cs),
        "pointsDepart": points_depart,
        "xp": xp,
        "reputations": "",
        "notes": "",
        "updatedAt": "2026-08-15T00:00:00.000Z",
    }

    height_raw = cs["AR6"].value
    try:
        character["heightM"] = float(str(height_raw).replace(",", "."))
    except (TypeError, ValueError):
        pass

    return character, warnings


# ---------------------------------------------------------------------------
# Écriture des fichiers de sortie
# ---------------------------------------------------------------------------

def write_reference_data_ts(ref: dict) -> None:
    out = ROOT / "src" / "shared" / "reference-data.ts"
    content = f"""// Fichier généré par scripts/import_xlsx.py — NE PAS ÉDITER À LA MAIN.
// Source : listes de "{REFERENCE_FILE.name}" (identiques dans les 8 fiches
// de Mon Drive/ADD40K/Fiches persos), croisées avec Règles ADD40K - V0.2.docx.
// Pour mettre à jour : relancer `npm run import:xlsx`.

import type {{ ReferenceData }} from "./types";

export const referenceData: ReferenceData = {json.dumps(
        {k: v for k, v in ref.items() if k != "localisations"}, ensure_ascii=False, indent=2
    )};

/** Table des localisations de blessure (char_sheet, identique sur toutes les fiches). */
export const LOCALISATIONS = {json.dumps(ref["localisations"], ensure_ascii=False)};
"""
    out.write_text(content, encoding="utf-8")
    note(f"écrit {out.relative_to(ROOT)}")


def write_characters_seed(characters: list[dict]) -> None:
    out = ROOT / "scripts" / "characters.seed.json"
    out.write_text(json.dumps(characters, ensure_ascii=False, indent=2), encoding="utf-8")
    note(f"écrit {out.relative_to(ROOT)} ({len(characters)} personnages)")


def main() -> None:
    if not SOURCE_DIR.is_dir():
        raise SystemExit(f"Dossier source introuvable : {SOURCE_DIR}")

    note("# Rapport d'import ADD40K\n")
    note(f"Source : `{SOURCE_DIR}`\n")

    ref = build_reference_data()
    write_reference_data_ts(ref)

    note("\n## Personnages\n")
    characters = []
    all_warnings: dict[str, list[str]] = {}
    for path in sorted(SOURCE_DIR.glob("*.xlsx")):
        character, warnings = extract_character(path, ref["advantages"])
        characters.append(character)
        all_warnings[character["name"]] = warnings
        status = "⚠️" if warnings else "✓"
        note(f"- {status} **{character['name']}** ({path.name}) — race: {character['race']}, "
             f"owner: {character['ownerUsername']}, solde à vérifier après calcul.")

    write_characters_seed(characters)

    note("\n## Anomalies détectées\n")
    any_warning = False
    for name, warnings in all_warnings.items():
        for w in warnings:
            any_warning = True
            note(f"- **{name}** : {w}")
    if not any_warning:
        note("Aucune.")

    note("\n## Rappels connus (voir plan lively-rolling-comet.md)\n")
    note("- Bug \"Dans les nuages\" (catalogue -30 vs +20 mis en cache par erreur dans le "
         "classeur d'origine) : corrigé automatiquement par cet import (les avantages sont "
         "dérivés du catalogue par libellé, pas de la cellule mise en cache par personnage). "
         "Le SENS du calcul du solde a aussi été corrigé côté app (calc-engine.ts) : un "
         "avantage coûte des points, un inconvénient en donne — confirmé par le MJ, contraire "
         "à la formule brute du classeur qui allait dans l'autre sens.")
    note("- gnome/orc : bonus d'Intelligence du docx (gnome +1, orc -1) non appliqué dans le "
         "classeur — laissé tel que joué, à confirmer avec le groupe si besoin.")
    note("- Stella a la race \"Illitide\", hors des 8 races gérées par le calculateur : bonus "
         "racial et points de compétence de départ à 0 pour l'instant.")
    note("- Les usernames des joueurs sont déduits des noms de fichiers Background quand "
         "possible (Karun→jo, Stern Tack→charles, Frigg→jack) ; les autres personnages ont un "
         "username provisoire (slug du nom) à corriger par le MJ après le premier déploiement.")

    report_path = ROOT / "scripts" / "import-report.md"
    report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    print(f"\nRapport complet : {report_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
